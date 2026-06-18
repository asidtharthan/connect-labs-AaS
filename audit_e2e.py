"""END-TO-END AUDIT — every number, traced raw->master->status->aggregates->payload,
reconciled against (1) the 10-Jun master baseline, (2) independent recompute, (3) the GW
workbook. Prints PASS/FAIL per check + a final accuracy summary. No spot checks.
"""
import csv as _csv
import json
import os
from collections import Counter, defaultdict
from datetime import date, timedelta

from openpyxl import load_workbook

import build_master_4src as bm  # the master under test

TODAY = date.today()  # match the build's dynamic time-gating for the independent status recompute
TOPICS = ["A", "B", "C", "D", "E", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
SG_ORDER = ["TRS", "TRE", "ABT1-A", "ABT1-B", "ABT2-A", "ABT2-B"]
ROLL = {"TRS": "TRS", "TRE": "TRE", "ABT1-A": "ABT1", "ABT1-B": "ABT1", "ABT2-A": "ABT2", "ABT2-B": "ABT2"}
results = []  # (section, check, passed, detail)


def chk(section, name, passed, detail=""):
    results.append((section, name, passed, detail))
    print(f"  [{'PASS' if passed else 'FAIL'}] {name}  {detail}")


print("=" * 90)
print("A. RAW SOURCE INTEGRITY")
print("=" * 90)
n_trig = sum(len(v) for v in bm.triggers_by_flw_iv.values())
n_wel = sum(len(v) for v in bm.welcome_flws_by_key.values())
sessions = json.loads(open("_ocs_state_cache.json").read())
n_tagged = sum(1 for s in sessions if s.get("pid") and s.get("interview") and str(s["interview"]).strip())
claimed_pairs = {
    (c, f) for c in bm.cohort_flws for f in bm.cohort_flws[c] if bm.cohort_flw_meta[(c, f)].get("date_claimed")
}
print(
    f"  triggers={n_trig}  welcome-keys={n_wel}  ocs_sessions={len(sessions)} tagged={n_tagged}  claimed_pairs={len(claimed_pairs)}"
)
# no duplicate trigger_form_id in master
fids = [r["trigger_form_id"] for r in bm.rows]
chk("A", "no duplicate trigger_form_id in master", len(fids) == len(set(fids)), f"{len(fids)} rows")
# every cohort maps to a subgroup
bad_sg = [c for c in bm.cohort_info if bm.cohort_to_sg(c) is None]
chk("A", "every cohort maps to a subgroup", not bad_sg, f"{len(bad_sg)} bad")

print("=" * 90)
print("B. MASTER vs 10-Jun BASELINE (row-level) + INVARIANTS")
print("=" * 90)
# Baseline comparison is OPTIONAL — the 10-Jun baseline holds participant ids and is not shipped
# server-side. When absent, the integrity invariants below still run (they need only bm.rows).
if os.path.exists("master_v7_2026-06-10.csv"):
    base = {r["trigger_form_id"]: r for r in _csv.DictReader(open("master_v7_2026-06-10.csv", encoding="utf-8"))}
    live = {r["trigger_form_id"]: r for r in bm.rows}
    shared = set(base) & set(live)
    only_base = set(base) - set(live)
    only_live = set(live) - set(base)
    # Growth-aware: live must be a SUPERSET of the 10-Jun baseline (new interviews expected as data
    # flows in). A regression = a baseline row that vanished from live. only_live>0 is fine (growth).
    chk(
        "B",
        "baseline rows all present in live (no coverage regression)",
        len(only_base) == 0,
        f"shared={len(shared)} only_live={len(only_live)} (growth) only_base={len(only_base)}",
    )
    struct_cols = [
        "cohort_id",
        "subgroup",
        "cohort_type",
        "interview_n",
        "topic_code",
        "topic_name",
        "training_date",
        "release_date",
    ]
    struct_mm = sum(1 for k in shared for c in struct_cols if str(live[k][c]) != str(base[k][c]))
    chk(
        "B",
        "structural columns bit-exact vs baseline",
        struct_mm == 0,
        f"{struct_mm} mismatches across {len(struct_cols)} cols",
    )
    st_reg = sum(1 for k in shared if base[k]["is_started"] == "Y" and live[k]["is_started"] == "N")
    co_reg = sum(1 for k in shared if base[k]["is_completed"] == "Y" and live[k]["is_completed"] == "N")
    st_fwd = sum(1 for k in shared if base[k]["is_started"] == "N" and live[k]["is_started"] == "Y")
    co_fwd = sum(1 for k in shared if base[k]["is_completed"] == "N" and live[k]["is_completed"] == "Y")
    chk("B", "is_started: zero regressions (Y->N)", st_reg == 0, f"regressions={st_reg}, forward N->Y={st_fwd}")
    chk("B", "is_completed: zero regressions (Y->N)", co_reg == 0, f"regressions={co_reg}, forward N->Y={co_fwd}")
else:
    print("  [SKIP] master_v7_2026-06-10 baseline not present — integrity invariants below still enforced.")
# invariants
inv1 = all(not (r["is_completed"] == "Y" and r["is_started"] != "Y") for r in bm.rows)
inv2 = all(not (r["is_started"] == "Y" and r["matched_session_id"] == "") for r in bm.rows)
sid_use = Counter(r["matched_session_id"] for r in bm.rows if r["matched_session_id"])
inv3 = all(v == 1 for v in sid_use.values())
inv4 = all(
    SUBGROUP_OK := (
        r["topic_code"] in bm.SUBGROUP_DESIGN[r["subgroup"]]["topics"]
        and bm.SUBGROUP_DESIGN[r["subgroup"]]["topics"].index(r["topic_code"]) + 1 == int(r["interview_n"])
    )
    for r in bm.rows
)
chk("B", "invariant completed=>started", inv1)
chk("B", "invariant started=>has session", inv2)
chk("B", "invariant no session double-claimed", inv3, f"{sum(1 for v in sid_use.values() if v>1)} dup sids")
chk("B", "invariant interview_n == position(topic) in subgroup (ALL rows)", inv4)

print("=" * 90)
print("C. STATUS TABLE (independent recompute, full grid)")
print("=" * 90)
mlook = {}


def rank(r):
    return (1 if r["is_completed"] == "Y" else 0) * 2 + (1 if r["is_started"] == "Y" else 0)


for r in bm.rows:
    k = (r["connect_id"], r["cohort_id"], r["topic_code"])
    if k not in mlook or rank(r) > rank(mlook[k]):
        mlook[k] = r


def status_for(flw, cohort, topic):
    sg = bm.cohort_to_sg(cohort)
    topics = bm.SUBGROUP_DESIGN[sg]["topics"]
    if topic not in topics:
        return "not-applicable"
    n = topics.index(topic) + 1
    m = mlook.get((flw, cohort, topic))
    if m and m["is_completed"] == "Y":
        return "completed"
    if m and m["is_started"] == "Y":
        return "started-not-completed"
    td = bm.cohort_info.get(cohort, {}).get("training_date")
    if not td:
        return "available-not-started"
    cad = bm.SUBGROUP_DESIGN[sg]["cadence"]
    if TODAY < td + timedelta(days=(n - 1) * cad):
        return "not-available-yet"
    if n < len(topics) and TODAY >= td + timedelta(days=n * cad):
        return "available-missed-overdue"
    return "available-not-started"


grid = {}
for cohort, flw in claimed_pairs:
    for topic in TOPICS:
        grid[(cohort, flw, topic)] = status_for(flw, cohort, topic)
chk(
    "C",
    "grid complete (claimed_pairs x 14, no dup/missing)",
    len(grid) == len(claimed_pairs) * 14,
    f"{len(grid)} == {len(claimed_pairs)*14}",
)
na = sum(1 for v in grid.values() if v == "not-applicable")
exp_na = sum(14 - len(bm.SUBGROUP_DESIGN[bm.cohort_to_sg(c)]["topics"]) for (c, f) in claimed_pairs)
chk("C", "not-applicable count exact", na == exp_na, f"{na} == {exp_na}")
# C: started/completed reconcile to master (claimed)
g_comp = sum(1 for v in grid.values() if v == "completed")
g_start = sum(1 for v in grid.values() if v in ("completed", "started-not-completed"))
m_comp = len(
    {
        (r["connect_id"], r["cohort_id"], r["topic_code"])
        for r in bm.rows
        if r["is_completed"] == "Y" and (r["cohort_id"], r["connect_id"]) in claimed_pairs
    }
)
m_start = len(
    {
        (r["connect_id"], r["cohort_id"], r["topic_code"])
        for r in bm.rows
        if r["is_started"] == "Y" and (r["cohort_id"], r["connect_id"]) in claimed_pairs
    }
)
chk("C", "status completed == master completed (claimed)", g_comp == m_comp, f"{g_comp} == {m_comp}")
chk("C", "status started == master started (claimed)", g_start == m_start, f"{g_start} == {m_start}")


# C: independent re-derivation of schedule states (second code path)
def status_v2(flw, cohort, topic):
    sg = bm.cohort_to_sg(cohort)
    topics = bm.SUBGROUP_DESIGN[sg]["topics"]
    if topic not in topics:
        return "not-applicable"
    n = topics.index(topic) + 1
    m = mlook.get((flw, cohort, topic))
    completed = bool(m) and m["is_completed"] == "Y"
    started = bool(m) and m["is_started"] == "Y"
    td = bm.cohort_info.get(cohort, {}).get("training_date")
    rel = (td + timedelta(days=(n - 1) * cad_)) if (cad_ := bm.SUBGROUP_DESIGN[sg]["cadence"]) and td else None
    nrel = (td + timedelta(days=n * cad_)) if td and n < len(topics) else None
    avail = rel is not None and TODAY >= rel
    overdue = nrel is not None and TODAY >= nrel
    if completed:
        return "completed"
    if started:
        return "started-not-completed"
    if td is None:
        return "available-not-started"
    if not avail:
        return "not-available-yet"
    return "available-missed-overdue" if overdue else "available-not-started"


mism = sum(1 for k, v in grid.items() if status_v2(k[1], k[0], k[2]) != v)
chk(
    "C",
    "status logic agrees across 2 independent code paths (ALL cells)",
    mism == 0,
    f"{mism} disagreements / {len(grid)}",
)
distinct_states = set(grid.values())
chk(
    "C",
    "every cell exactly one of the 6 states",
    distinct_states
    <= set(
        bm.__dict__.get("STATES", [])
        or [
            "not-applicable",
            "not-available-yet",
            "available-not-started",
            "available-missed-overdue",
            "started-not-completed",
            "completed",
        ]
    ),
    f"states seen: {sorted(distinct_states)}",
)

print("=" * 90)
print("D. AGGREGATES — payload vs INDEPENDENT recompute (every cell)")
print("=" * 90)
payload = json.loads(open("payload_agg.json", encoding="utf-8").read())
# independent funnel recompute
cell = {}
for r in bm.rows:
    k = (r["connect_id"], r["cohort_id"], int(r["interview_n"]))
    c = cell.setdefault(
        k,
        {"sg": r["subgroup"], "n": int(r["interview_n"]), "flw": r["connect_id"], "t": False, "s": False, "c": False},
    )
    c["t"] = True
    if r["is_started"] == "Y":
        c["s"] = True
    if r["is_completed"] == "Y":
        c["c"] = True
cells = list(cell.values())
elig_sg = defaultdict(set)
for (cohort, topic), flws in bm.welcome_flws_by_key.items():
    sg = bm.cohort_to_sg(cohort)
    if sg:
        elig_sg[sg] |= flws
fre = defaultdict(lambda: {"t": set(), "s": set(), "c": set()})
for c in cells:
    f = fre[(c["sg"], c["n"])]
    if c["t"]:
        f["t"].add(c["flw"])
    if c["s"]:
        f["s"].add(c["flw"])
    if c["c"]:
        f["c"].add(c["flw"])
fmm = 0
for row in payload["funnel"]:
    f = fre[(row["sg"], row["n"])]
    if (
        len(f["t"]) != row["trig"]
        or len(f["s"]) != row["started"]
        or len(f["c"]) != row["completed"]
        or len(elig_sg[row["sg"]]) != row["elig"]
    ):
        fmm += 1
chk("D", "funnel: payload == independent recompute (all 25 rows)", fmm == 0, f"{fmm} mismatched rows")


# Tables independent
def agg(keyfn, keys):
    a = defaultdict(lambda: {"flw": set(), "ist": 0, "icmp": 0})
    for c in cells:
        for k in set(keyfn(c)):
            if c["s"]:
                a[k]["flw"].add(c["flw"])
                a[k]["ist"] += 1
            if c["c"]:
                a[k]["icmp"] += 1
    return a


t1a = agg(lambda c: [ROLL[c["sg"]], "Overall"], None)
t1mm = sum(
    1
    for row in payload["table1"]
    if not (
        len(t1a[row["key"]]["flw"]) == row["flws"]
        and t1a[row["key"]]["ist"] == row["ist"]
        and t1a[row["key"]]["icmp"] == row["icmp"]
    )
)
chk("D", "Table1: payload == recompute", t1mm == 0, f"{t1mm} mismatched")
t3a = agg(lambda c: ([c["sg"], "Overall"] if c["sg"].startswith(("ABT1", "ABT2")) else []), None)
t3mm = sum(
    1
    for row in payload["table3"]
    if not (
        len(t3a[row["key"]]["flw"]) == row["flws"]
        and t3a[row["key"]]["ist"] == row["ist"]
        and t3a[row["key"]]["icmp"] == row["icmp"]
    )
)
chk("D", "Table3: payload == recompute", t3mm == 0, f"{t3mm} mismatched")
t2a = defaultdict(lambda: {"flw": set(), "ist": 0, "icmp": 0})
for c in cells:
    tc = bm.SUBGROUP_DESIGN[c["sg"]]["topics"][c["n"] - 1]
    if c["s"]:
        t2a[tc]["flw"].add(c["flw"])
        t2a[tc]["ist"] += 1
    if c["c"]:
        t2a[tc]["icmp"] += 1
t2mm = sum(
    1
    for row in payload["table2"]
    if not (
        len(t2a[row["code"]]["flw"]) == row["flws"]
        and t2a[row["code"]]["ist"] == row["ist"]
        and t2a[row["code"]]["icmp"] == row["icmp"]
    )
)
chk("D", "Table2: payload == recompute", t2mm == 0, f"{t2mm} mismatched")
# topic_status dist vs grid
tsmm = 0
for row in payload["topic_status"]:
    for s in payload["states"]:
        if sum(1 for k, v in grid.items() if k[2] == row["code"] and v == s) != row[s]:
            tsmm += 1
chk("D", "topic_status dist: payload == grid recompute (14x6 cells)", tsmm == 0, f"{tsmm} mismatched cells")
# line series == funnel pct_started
lmm = 0
for sg in SG_ORDER:
    fr = [r for r in payload["funnel"] if r["sg"] == sg]
    if [r["pct_started"] for r in sorted(fr, key=lambda x: x["n"])] != payload["line_pct_started"][sg]:
        lmm += 1
chk("D", "line series == funnel %started", lmm == 0, f"{lmm} subgroups mismatched")

print("=" * 90)
print("E. CROSS-CONSISTENCY")
print("=" * 90)
# Table2.ist == sum funnel.started over (sg,n) with that topic  (one FLW=one cohort per subgroup => unique==count)
ok = True
for row in payload["table2"]:
    s = sum(r["started"] for r in payload["funnel"] if r["topic"] == row["code"])
    if s != row["ist"]:
        ok = False
chk("E", "Table2 interviews-started == Σ funnel started by topic", ok)
# funnel monotonic: completed<=started<=triggered<=eligible
mono = all(r["completed"] <= r["started"] <= r["trig"] <= r["elig"] for r in payload["funnel"])
chk("E", "funnel monotonic completed<=started<=triggered<=eligible", mono)
# Overall FLWs (table1) == unique started connect_ids (dedup across subgroups)
overall_flw = len({c["flw"] for c in cells if c["s"]})
t1_overall = next(r["flws"] for r in payload["table1"] if r["key"] == "Overall")
chk(
    "E",
    "Table1 Overall FLWs == dedup unique started FLWs (no inflation)",
    overall_flw == t1_overall,
    f"{overall_flw} == {t1_overall}",
)

print("=" * 90)
print("F. GW WORKBOOK RECONCILIATION (11-Jun reference; live is newer -> forward drift)")
print("=" * 90)
_GW_XLSX = "screenshots/Latest files/GW Tables - 11th June 2026.xlsx"
if not os.path.exists(_GW_XLSX):
    print(
        f"  [SKIP] GW reference workbook not present ({_GW_XLSX}) — one-time reference check, "
        f"not a data-integrity gate. Other checks still enforced."
    )
else:
    wb = load_workbook(_GW_XLSX, read_only=True, data_only=True)
    # parse Retention Drop-off funnel
    ws = wb["Retention Drop-off"]
    rows_ws = [list(r) for r in ws.iter_rows(values_only=True)]
    gw_funnel = {}
    for r in rows_ws:
        if r and r[0] in SG_ORDER:
            sg = r[0]
            # interview blocks start at col 9, step 8: Topic,Elig,Trig,%,Start,%,Compl,%
            for n in range(1, 9):
                base_c = 8 + (n - 1) * 8 + 1
                topic = r[base_c] if base_c < len(r) else None
                if topic in (None, "-", ""):
                    continue
                try:
                    elig = int(r[base_c + 1])
                    trig = int(r[base_c + 2])
                    start = int(r[base_c + 4])
                    comp = int(r[base_c + 6])
                except (TypeError, ValueError):
                    continue
                gw_funnel[(sg, n)] = (elig, trig, start, comp)
    exact = drift_fwd = drift_other = 0
    for row in payload["funnel"]:
        key = (row["sg"], row["n"])
        if key not in gw_funnel:
            continue
        ge, gt, gs, gc = gw_funnel[key]
        le, lt, ls, lc = row["elig"], row["trig"], row["started"], row["completed"]
        if (ge, gt, gs, gc) == (le, lt, ls, lc):
            exact += 1
        elif le >= ge and lt >= gt and ls >= gs and lc >= gc:
            drift_fwd += 1
        else:
            drift_other += 1
            print(f"    OTHER-DRIFT {key}: live(e/t/s/c)={le}/{lt}/{ls}/{lc} gw={ge}/{gt}/{gs}/{gc}")
    chk(
        "F",
        "funnel vs GW: all cells exact OR forward-drift (live>=gw)",
        drift_other == 0,
        f"exact={exact} forward-drift={drift_fwd} other={drift_other}",
    )

print("\n" + "=" * 90)
print("AUDIT SUMMARY")
print("=" * 90)
passed = sum(1 for _, _, p, _ in results if p)
for sec in ["A", "B", "C", "D", "E", "F"]:
    secr = [r for r in results if r[0] == sec]
    print(f"  Section {sec}: {sum(1 for r in secr if r[2])}/{len(secr)} passed")
print(f"\n  TOTAL: {passed}/{len(results)} checks passed")
print(
    "  RESULT:", "ALL PASS — 200% reconciled" if passed == len(results) else f"*** {len(results)-passed} FAILURES ***"
)
