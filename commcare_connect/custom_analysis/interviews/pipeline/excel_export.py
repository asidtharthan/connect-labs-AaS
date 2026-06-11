"""
Generate the answers.xlsx file from InterviewAnswer rows.

Schema mirrors Neal's R Step 1 output (same column order). Streamed in-memory
via openpyxl so we don't need S3.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Column order — matches the current step1.r output schema (13 columns).
ANSWERS_COLUMNS = [
    "session_id",
    "participant_id",
    "qid",
    "parent_qid",
    "question_text",
    "topic",
    "answer_text",
    "asked",
    "completed",
    "num_turns",
    "word_count",
    "time_to_answer_seconds",
    "extraction_method",
]

# Reasonable column widths (in characters) for readability when opened.
COLUMN_WIDTHS = {
    "session_id": 38,
    "participant_id": 22,
    "qid": 8,
    "parent_qid": 10,
    "question_text": 45,
    "topic": 22,
    "answer_text": 60,
    "asked": 7,
    "completed": 10,
    "num_turns": 10,
    "word_count": 10,
    "time_to_answer_seconds": 18,
    "extraction_method": 18,
}


def build_answers_xlsx(rows: list[dict]) -> bytes:
    """Build the answers.xlsx file in memory and return its bytes.

    Args:
        rows: list of dicts matching the InterviewAnswer schema.

    Returns:
        bytes of a complete xlsx file (ready to stream as HttpResponse).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "answers"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, col_name in enumerate(ANSWERS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Set column widths
    for col_idx, col_name in enumerate(ANSWERS_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 15)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(ANSWERS_COLUMNS, start=1):
            value = row.get(col_name)
            # Excel can't render None for booleans cleanly; convert to actual bools/blanks
            if isinstance(value, bool):
                ws.cell(row=row_idx, column=col_idx, value=value)
            elif value is None:
                ws.cell(row=row_idx, column=col_idx, value="")
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Stream to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# Per-cohort table XLSX — same columns as the dashboard table. Used by the
# "Download cohort table" button so Ali can run her A/B comparison in Excel.
COHORT_COLUMNS = [
    ("cohort", "Cohort"),
    ("opportunity_id", "Opportunity ID"),
    ("flws_claimed", "FLWs Claimed Job"),
    ("flws_completed_assessment", "FLWs Completed Assessment"),
    ("flws_with_completed_interview", "FLWs with >=1 Interview"),
    ("connect_services_delivered", "Services Delivered (Connect)"),
    ("connect_approved", "Connect Approved"),
    ("connect_over_limit", "Connect Over Limit"),
    ("interviews_completed", "Total Interviews (OCS)"),
    ("total_expected_interviews", "Expected Interviews"),
    ("completion_rate", "Completion Rate (%)"),
    ("median_message_length", "Median Word Count"),
]

COHORT_COLUMN_WIDTHS = {
    "cohort": 32,
    "opportunity_id": 14,
    "flws_claimed": 16,
    "flws_completed_assessment": 26,
    "flws_with_completed_interview": 22,
    "connect_services_delivered": 26,
    "connect_approved": 18,
    "connect_over_limit": 18,
    "interviews_completed": 22,
    "total_expected_interviews": 20,
    "completion_rate": 16,
    "median_message_length": 18,
}


def build_cohort_xlsx(rows: list[dict], totals: dict | None = None) -> bytes:
    """Build a cohort-rollup xlsx mirroring the dashboard's per-cohort table.

    Args:
        rows: list of cohort dicts (from `InterviewsDataAccess.get_per_cohort_metrics`).
        totals: optional totals dict; if provided, written as a footer row.

    Returns:
        Bytes of a complete xlsx (one sheet, "cohort_metrics").
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "cohort_metrics"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, (_, label) in enumerate(COHORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for col_idx, (key, _) in enumerate(COHORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COHORT_COLUMN_WIDTHS.get(key, 15)

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(COHORT_COLUMNS, start=1):
            value = row.get(key)
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    if totals:
        footer_row = len(rows) + 3
        bold = Font(bold=True)
        ws.cell(row=footer_row, column=1, value="TOTALS").font = bold
        # Map totals keys to columns
        total_map = {
            "flws_claimed": totals.get("total_flws_claimed", 0),
            "flws_completed_assessment": totals.get("total_flws_completed_assessment", 0),
            "flws_with_completed_interview": totals.get("total_flws_with_completed_interview", 0),
            "connect_services_delivered": totals.get("total_connect_services_delivered", 0),
            "interviews_completed": totals.get("total_interviews_completed", 0),
            "total_expected_interviews": totals.get("total_expected_interviews", 0),
            "completion_rate": totals.get("overall_completion_rate", 0),
        }
        for col_idx, (key, _) in enumerate(COHORT_COLUMNS, start=1):
            if key in total_map:
                cell = ws.cell(row=footer_row, column=col_idx, value=total_map[key])
                cell.font = bold

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
